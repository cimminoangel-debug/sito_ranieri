import uuid
import io
import os
import qrcode
import openpyxl
import stripe
from django.utils import timezone # <-- Aggiungi questa importazione in cima
from django.contrib.auth.decorators import user_passes_test # <-- Aggiungi in cima
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.conf import settings
from django.core.mail import EmailMessage 
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from .models import Canzone, Concerto, Biografia, Partecipante, VideoSpettacolo

def controllo_staff(user):
    return user.is_authenticated and user.is_staff

# Viste Principali del Sito
def home(request):
    oggi = timezone.now().date()
    concerti_futuri = Concerto.objects.filter(data__gte=oggi).order_by('data')
    ultimi_video = VideoSpettacolo.objects.all().order_by('-data_aggiunta')[:2]
    
    return render(request, 'tributo/home.html', {
        'concerti': concerti_futuri,
        'video_spettacoli': ultimi_video
    })


# NUOVA VISTA: Elenco completo di tutti i concerti (Archivio)
def tutti_i_concerti(request):
    # Prende tutti i concerti ordinati dal più recente al più vecchio
    concerti_totali = Concerto.objects.all().order_by('-data')
    return render(request, 'tributo/tutti_i_concerti.html', {'concerti': concerti_totali})

def biografia(request):
    lista_biografie = Biografia.objects.all().order_by('anno_inizio_carriera')
    return render(request, 'tributo/biografia.html', {'biografie': lista_biografie})

def discografia(request):
    canzoni_reali = Canzone.objects.all().order_by('anno_uscita')
    return render(request, 'tributo/discografia.html', {'canzoni': canzoni_reali})


# Acquisto delle Canzoni
def ordine_riepilogo(request, canzone_id):
    canzone = get_object_or_404(Canzone, id=canzone_id)
    
    if request.method == "POST":
        quantita = int(request.POST.get('quantita', 1))
        nome_acquirente = request.POST.get('nome_completo', 'Acquirente') # Adatta se i tuoi campi hanno nomi diversi
        email_acquirente = request.POST.get('email', '')

        # Memorizziamo temporaneamente i dettagli dell'acquisto nella sessione di Django
        request.session['dati_acquisto_canzone'] = {
            'canzone_id': canzone.id,
            'quantita': quantita,
            'nome': nome_acquirente,
            'email': email_acquirente
        }

        # Convertiamo il prezzo in centesimi per Stripe
        prezzo_in_centesimi = int(canzone.prezzo * 100)

        # Chiamata API a Stripe per la sessione di pagamento della musica
        sessione_checkout = stripe.checkout.Session.create(
            payment_method_types=['card'],
            line_items=[{
                'price_data': {
                    'currency': 'eur',
                    'product_data': {
                        'name': f"Download Digitale: {canzone.titolo}",
                        'description': f"Successo discografico del {canzone.anno_uscita}",
                    },
                    'unit_amount': prezzo_in_centesimi,
                },
                'quantity': quantita,
            }],
            mode='payment',
            success_url=request.build_absolute_uri('/ordine/canzone/pagamento/successo/'),
            cancel_url=request.build_absolute_uri(f'/ordine/{canzone.id}/riepilogo/'),
        )
        return redirect(sessione_checkout.url, code=303)

    return render(request, 'tributo/ordine_riepilogo.html', {'canzone': canzone})


def canzone_pagamento_successo(request):
    dati = request.session.get('dati_acquisto_canzone')
    
    if not dati:
        return redirect('discografia')
        
    canzone = get_object_or_404(Canzone, id=dati['canzone_id'])
    transazione_id = str(uuid.uuid4()).split('-')[0].upper()
    totale = canzone.prezzo * dati['quantita']
    
    # Didattica: Simulazione invio email per la musica nel terminale di Django
    print("\n" + "="*50)
    print(f"INVIO RICEVUTA MUSICALE A: {dati['email']}")
    print(f"OGGETTO: Il tuo download per {canzone.titolo}")
    print(f"TESTO: Grazie {dati['nome']}. Codice Transazione: #TRS-{transazione_id}")
    print("="*50 + "\n")
    
    # Puliamo la sessione temporanea
    del request.session['dati_acquisto_canzone']
    
    return render(request, 'tributo/ordine_transazione.html', {
        'canzone': canzone,
        'transazione_id': transazione_id,
        'quantita': dati['quantita'],
        'totale': totale,
        'nome': dati['nome'],
        'email': dati['email']
    })


# Acquisto dei Biglietti del Concerto
stripe.api_key = settings.STRIPE_SECRET_KEY
def concerto_riepilogo(request, concerto_id):
    concerto = get_object_or_404(Concerto, id=concerto_id)
    return render(request, 'tributo/concerto_riepilogo.html', {'concerto': concerto})

def concerto_riepilogo(request, concerto_id):
    concerto = get_object_or_404(Concerto, id=concerto_id)
    
    if request.method == "POST":
        quantita = int(request.POST.get('quantita', 1))
        
        # Recuperiamo i nomi degli intestatari inseriti nel form
        nomi_biglietti = request.POST.getlist('nomi_biglietti')
        email_biglietti = request.POST.getlist('email_biglietti')
        
        # Salviamo temporaneamente i dati nella sessione di Django. 
        # Verranno inseriti nel database reale SOLO se il pagamento va a buon fine.
        request.session['dati_acquisto_concerto'] = {
            'concerto_id': concerto.id,
            'quantita': quantita,
            'nomi': nomi_biglietti,
            'email': email_biglietti
        }

        # Convertiamo il prezzo in centesimi (Stripe richiede i centesimi, es: €25.00 diventa 2500)
        prezzo_in_centesimi = int(concerto.prezzo * 100)

        # Creazione della sessione di pagamento sul server di Stripe
        sessione_checkout = stripe.checkout.Session.create(
            payment_method_types=['card'], # Accetta carte di credito, Apple Pay, Google Pay
            line_items=[{
                'price_data': {
                    'currency': 'eur',
                    'product_data': {
                        'name': f"Biglietto Spettacolo: {concerto.luogo}",
                        'description': f"{concerto.citta} — {concerto.data}",
                    },
                    'unit_amount': prezzo_in_centesimi,
                },
                'quantity': quantita,
            }],
            mode='payment',
            # Se il pagamento riesce, Stripe rimanda l'utente a questo indirizzo
            success_url=request.build_absolute_uri('/concerto/pagamento/successo/'),
            # Se l'utente annulla, viene rimandato qui
            cancel_url=request.build_absolute_uri(f'/concerto/{concerto.id}/riepilogo/'),
        )

        # Reindirizziamo l'utente direttamente sul portale di pagamento sicuro di Stripe
        return redirect(sessione_checkout.url, code=303)

    return render(request, 'tributo/concerto_riepilogo.html', {'concerto': concerto})
    send_mail(subject, message, from_email, recipient_list, fail_silently=False)



    nomi_partecipanti = request.GET.getlist('nomi_biglietti')
    email_partecipanti = request.GET.getlist('email_biglietti')
    
    transazione_id = str(uuid.uuid4()).split('-')[0].upper()
    biglietti_salvati = []

    for i in range(len(nomi_partecipanti)):
        nome = nomi_partecipanti[i]
        email = email_partecipanti[i]
        codice_unico = f"TKT-{transazione_id}-0{i+1}"
        
        nuovo_p = Partecipante.objects.create(
            concerto=concerto,
            nome_completo=nome,
            email=email,
            codice_biglietto=codice_unico
        )
        biglietti_salvati.append(nuovo_p)
        invia_email_con_biglietto(nuovo_p)

    totale = concerto.prezzo * quantita
    
    return render(request, 'tributo/concerto_transazione.html', {
        'concerto': concerto,
        'totale': totale,
        'transazione_id': transazione_id,
        'partecipanti': biglietti_salvati
    })


# --- MOTORE DI GENERAZIONE GENERALE DEL PDF ---
def genera_pdf_biglietto(partecipante):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    
    p.setFillColorRGB(0.29, 0.08, 0.15)
    p.rect(50, 700, 500, 50, fill=True, stroke=False)
    
    p.setFillColorRGB(1, 1, 1)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(70, 718, "BIGLIETTO DIGITALE - EVENTO UFFICIALE")
    
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, 650, f"Spettacolo: {partecipante.concerto.luogo}")
    p.setFont("Helvetica", 12)
    p.drawString(50, 630, f"Citta e data: {partecipante.concerto.citta} -- {partecipante.concerto.data}")
    
    p.setStrokeColorRGB(0.8, 0.8, 0.8)
    p.line(50, 610, 550, 610)
    
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, 580, f"Titolare: {partecipante.nome_completo}")
    p.setFont("Helvetica", 11)
    p.drawString(50, 560, f"Email: {partecipante.email}")
    p.drawString(50, 540, f"Codice Univoco: {partecipante.codice_biglietto}")
    
    # QR Code di rete puntato all'IP locale del server Debian
    link_validazione = f"http://192.168.0.101:8000/gestione/valida/{partecipante.codice_biglietto}/"
    
    qr = qrcode.QRCode(version=1, box_size=10, border=1)
    qr.add_data(link_validazione)
    qr.make(fit=True)
    
    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_path = os.path.join(settings.BASE_DIR, 'media-serve', f'temp_qr_{partecipante.id}.png')
    qr_img.save(qr_path)
    
    p.drawImage(qr_path, 200, 340, width=200, height=200)
    
    p.setFont("Helvetica-Oblique", 9)
    p.setFillColorRGB(0.5, 0.5, 0.5)
    p.drawString(50, 290, "* Mostra questo QR Code dal tuo smartphone all'ingresso del teatro per la convalida del posto.")
    
    p.showPage()
    p.save()
    
    if os.path.exists(qr_path):
        os.remove(qr_path)
        
    buffer.seek(0)
    return buffer.getvalue()


def invia_email_con_biglietto(partecipante):
    pdf_data = genera_pdf_biglietto(partecipante)
    email = EmailMessage(
        subject=f"Il tuo Biglietto per {partecipante.concerto.luogo}",
        body=f"Gentile {partecipante.nome_completo},\n\nin allegato trovi il tuo biglietto ufficiale in formato PDF.\n\nCi vediamo allo spettacolo!\nStaff Artista Ufficiale",

        from_email='biglietteria@sitotributo.it',
        to=[partecipante.email],
    )
    email.attach(f"Biglietto_{partecipante.codice_biglietto}.pdf", pdf_data, "application/pdf")
    email.send()


# === FUNZIONE DI APERTURA E VISUALIZZAZIONE DEL PDF SULLA DASHBOARD ===
def vedi_pdf_biglietto(request, partecipante_id):
    partecipante = get_object_or_404(Partecipante, id=partecipante_id)
    pdf_data = genera_pdf_biglietto(partecipante)
    
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Biglietto_{partecipante.codice_biglietto}.pdf"'
    return response


# Validazione Varchi e Controllo Accessi
@user_passes_test(controllo_staff, login_url='/admin/login/') # <-- Aggiungi questa riga
def valida_biglietto(request, codice_biglietto):
    partecipante = get_object_or_404(Partecipante, codice_biglietto=codice_biglietto)
    stato_precedente = partecipante.utilizzato
    errore_duplicato = False

    if stato_precedente:
        errore_duplicato = True
    else:
        partecipante.utilizzato = True
        partecipante.save()

    return render(request, 'tributo/valida_biglietto.html', {
        'p': partecipante,
        'errore_duplicato': errore_duplicato
    })


# Dashboard Registro Biglietti ed Eventi
@user_passes_test(controllo_staff, login_url='/admin/login/') # <-- Aggiungi questa riga
def lista_partecipanti(request):
    # 1. Recupera i parametri inviati dal modulo HTML
    concerto_filtrato_id = request.GET.get('concerto_id')
    testo_ricerca = request.GET.get('q', '') # Raccoglie la stringa di testo cercata
    
    tutti_i_concerti = Concerto.objects.all()
    
    # 2. Base di partenza: tutti i partecipanti
    partecipanti = Partecipante.objects.all().order_by('-data_acquisto')
    
    # 3. Applica il filtro del concerto (se selezionato)
    if concerto_filtrato_id:
        partecipanti = partecipanti.filter(concerto_id=concerto_filtrato_id)
        
    # 4. Applica il filtro testuale sul nome/cognome (se digitato)
    if testo_ricerca:
        # __icontains cerca all'interno della stringa ignorando maiuscole/minuscole
        partecipanti = partecipanti.filter(nome_completo__icontains=testo_ricerca)
        
    # 5. Ricalcola i contatori dinamici basandoti sui risultati filtrati
    totale_venduti = partecipanti.count()
    totale_ingressi = partecipanti.filter(utilizzato=True).count()
        
    return render(request, 'tributo/lista_partecipanti.html', {
        'partecipanti': partecipanti,
        'concerti': tutti_i_concerti,
        'concerto_selezionato_id': concerto_filtrato_id,
        'testo_ricerca': testo_ricerca,        # <-- Passiamo il testo attuale al template
        'totale_venduti': totale_venduti,     
        'totale_ingressi': totale_ingressi   
    })


@user_passes_test(controllo_staff, login_url='/admin/login/') # <-- Aggiungi questa riga
def reinvia_biglietto(request, partecipante_id):
    partecipante = get_object_or_404(Partecipante, id=partecipante_id)
    invia_email_con_biglietto(partecipante)
    messages.success(request, f"Biglietto in formato PDF reinviato con successo a {partecipante.email}!")
    return redirect('lista_partecipanti')

@user_passes_test(controllo_staff, login_url='/admin/login/') # <-- Aggiungi questa riga
def esporta_excel_partecipanti(request):
    # 1. Recupera gli stessi filtri attivi nella dashboard
    concerto_filtrato_id = request.GET.get('concerto_id')
    testo_ricerca = request.GET.get('q', '')
    
    # 2. Applica i filtri alla query del database
    partecipanti = Partecipante.objects.all().order_by('-data_acquisto')
    if concerto_filtrato_id:
        partecipanti = partecipanti.filter(concerto_id=concerto_filtrato_id)
    if testo_ricerca:
        partecipanti = partecipanti.filter(nome_completo__icontains=testo_ricerca)
        
    # 3. Crea il documento Excel in memoria usando openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro Biglietti"
    
    # 4. Scrive la riga di intestazione (Header)
    intestazioni = ["Codice Biglietto", "Nominativo", "Email", "Spettacolo / Luogo", "Data Acquisto", "Stato Ingressi"]
    ws.append(intestazioni)
    
    # Rende l'intestazione in grassetto (Stile grafico opzionale)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    # 5. Popola le righe con i dati reali estratti da SQLite
    for p in partecipanti:
        stato_testo = "🚫 Entrato / Usato" if p.utilizzato else "🎟️ Attivo / Libero"
        data_formattata = p.data_acquisto.strftime("%d/%m/%Y %H:%M")
        
        riga = [
            p.codice_biglietto,
            p.nome_completo,
            p.email,
            f"{p.concerto.luogo} ({p.concerto.citta})",
            data_formattata,
            stato_testo
        ]
        ws.append(riga)
        
    # Autoadatta la larghezza delle colonne in base al testo contenuto
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 6. Prepara la risposta HTTP dicendo al browser che si tratta di un file Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Registro_Partecipanti_Teatro.xlsx"'
    
    # Salva il file direttamente nello stream di risposta
    wb.save(response)
    return response

def concerto_pagamento_successo(request):
    # Recuperiamo i dati dell'ordine lasciati in sospeso nella sessione
    dati = request.session.get('dati_acquisto_concerto')
    
    if not dati:
        # Se non ci sono dati in sessione, reindirizza alla home per sicurezza
        return redirect('home')
        
    concerto = get_object_or_404(Concerto, id=dati['concerto_id'])
    transazione_id = str(uuid.uuid4()).split('-')[0].upper()
    biglietti_salvati = []

    # Ora che il pagamento è REALE, salviamo i partecipanti nel DB
    for i in range(len(dati['nomi'])):
        nome = dati['nomi'][i]
        email = dati['email'][i]
        codice_unico = f"TKT-{transazione_id}-0{i+1}"
        
        nuovo_p = Partecipante.objects.create(
            concerto=concerto,
            nome_completo=nome,
            email=email,
            codice_biglietto=codice_unico
        )
        biglietti_salvati.append(nuovo_p)
        invia_email_con_biglietto(nuovo_p) # Invia l'email con il PDF reale

    # Puliamo la sessione temporanea per liberare la memoria
    del request.session['dati_acquisto_concerto']
    
    totale = concerto.prezzo * dati['quantita']

    return render(request, 'tributo/concerto_transazione.html', {
        'concerto': concerto,
        'totale': totale,
        'transazione_id': transazione_id,
        'partecipanti': biglietti_salvati
    })
@user_passes_test(controllo_staff, login_url='/admin/login/')
def dashboard_scanner(request):
    # Questa pagina serve solo a caricare l'interfaccia della fotocamera
    return render(request, 'tributo/dashboard_scanner.html')
